#!/usr/bin/python3
import openpyxl
from openpyxl.utils import column_index_from_string

'''
Segundo a ser executado
Gerará a saida na tela, podendo sem salvo em um arquivo csv usando o redirecionador >
'''

# ####### Variavies de configuracoes #######

arqA = "Aceitacao-Dependencias.xlsx"  # Arquivo com relação de srv aplicação
arqB = "DC-Migation-De-Para.xlsx"  # Arquivo com todos os srv

abaA = "Consolidado"  # Aba que contem o grupo de dados maior
abaB = "Total das Ondas"

# Numero da linha que os dados iniciam
inicioA = 2  
inicioB = 8 

# Relacionado ao arquivo A
# 0srcip, 1ambiente, 2dstip, 3app, 4porta
colunasA = ["B", "C", "E", "K", "I"]

# Relacionado ao arquivo B
colips = "U"  # Colna para comparacao
colTarget = "E"  # Coluna com o item desejado

# ##########################################

def main():
    # Abrindo arquivos
    wbA = openpyxl.load_workbook(arqA, data_only=True)
    wbB = openpyxl.load_workbook(arqB)
    
    # Carregando abas
    sheetA = wbA.get_sheet_by_name(abaA)
    sheetB = wbB.get_sheet_by_name(abaB)
    
    srcipB = []  # Lista para armazenar os produtos e pegar os codigos
    target = []  # Codigos dos repectivos produtos

    # Convertendo letras de colunas em numeros
    colunasAn = []  # Listar para armazenar valores 
    for letra in colunasA:
        colunasAn.append(column_index_from_string(letra))

    # Armazenados valores da planilha B
    for row in range(inicioB, sheetB.max_row + 1):
        tmp = str(sheetB[colips + str(row)].value)
        tmpnew = tmp.strip().replace(" ", "")  # Removendo espacos em branco
        srcipB.append(tmpnew) # Adicionar a lista
        target.append(sheetB[colTarget + str(row)].value)

    # Recebendo o primeiro codigo de aplicativo
    codapp = sheetA.cell(row=2, column=colunasAn[3]).value
    ipstemp = []  # Lista vazia para todos os ips da planilha A
    
    # Correndo toda a planilha para identificar quais IPs pertence as aplicacoes
    for row in range(inicioA, sheetA.max_row + 1):
        # Caso o codapp seja diferente do codigo da aplicacao 
        if codapp != str(sheetA.cell(row=row, column=colunasAn[3]).value):
            # Verificando se a lista de aplicacao esta vazio para ser pulado
            if len(ipstemp) != 0:
                ipsunicos = list(dict.fromkeys(ipstemp)) # Removendo elementos duplicados

                # Correndo todos os ips da lista 
                for ip in ipsunicos:
                    # Verificando se ip existe dentro da lista
                    if ip in srcipB:
                        index = srcipB.index(ip)
                        print (codapp + "," + ip + "," + target[index])
                    else:
                        print (codapp + "," + ip + "," + "hostname nao encontrado")

            #else: # Mostrar aplicacoes vazias
            #    print ("lista vazia", codapp)
                   
            # Armazenando o novo codigo da aplicacao
            codapp = sheetA.cell(row=row, column=colunasAn[3]).value
            # Limpar a lista dos ips temporario
            ipstemp = []

        # Caso o codapp seja igual, acumular srcips e dstips
        if codapp == str(sheetA.cell(row=row, column=colunasAn[3]).value): 
            # Caso seja diferente de aceitacao pulara a linha
            if str(sheetA.cell(row=row, column=colunasAn[1]).value) != "Aceitação":
                continue
            
            # Caso seja porta 445 ou 53 pulara a linha
            porta = str(sheetA.cell(row=row, column=colunasAn[4]).value)
            if porta == "445" or porta == "53":
                continue
            
            # Armazenando na lista ips de origem e destino
            ipstemp.append(sheetA.cell(row=row, column=colunasAn[0]).value)
            ipstemp.append(sheetA.cell(row=row, column=colunasAn[2]).value)


if __name__ == '__main__':
    main()
