#!/usr/bin/python3
import openpyxl
from openpyxl.utils import column_index_from_string

'''
Primeiro script a ser executado
Coloca apenas o codigo do aplicaitov na coluna K
'''

# ####### Variavies de configuracoes #######

arqA = "Aceitacao-Dependencias.xlsx"  # Arquivo com relação de srv aplicação
arqB = "DC-Migation-De-Para.xlsx"  # Arquivo com todos os srv

abaA = "Consolidado"  # Aba que contem o grupo de dados maior
colMatchA = "A"  # Coluna que os dados seram comparados com abaB
inicioA = 2  # Numero da linha que os dados iniciam

colapp = "K"

# ##########################################

wbA = openpyxl.load_workbook(arqA)
sheetA = wbA.get_sheet_by_name(abaA)

# Transformando coluna letra em numero
colappnova = column_index_from_string(colapp)

# Percorrendo aba com servidores
for row in range(inicioA, sheetA.max_row + 1):
    srvtemp = sheetA[colMatchA + str(row)].value
    if "-" in srvtemp:
        appcod = srvtemp[7:10]
    else:
        appcod = srvtemp[4:7]
    sheetA.cell(row=row, column=colappnova).value = appcod

wbA.save(arqA) # Savando o arquivo




