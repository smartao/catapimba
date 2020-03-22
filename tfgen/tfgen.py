#!/usr/bin/python3
import openpyxl
from openpyxl.utils import column_index_from_string

# ######### Variavies de configuracoes #########

arqA = "regras.xlsx"
abaA = "Regras"
inicioA = 6  # Inicio
fimA = 31
# Colunas com dados a serem processados
# nome, srcaddress, dstport, destaddress, protocol
colunasA = ["C", "G", "J", "I", "E"]

esp = "  "  # Espacamento padrao terraforme


def proc_string(string, dot):
    # Definindo o delimitador da string
    # dot = True significa que o delimitador pode ser o ponto
    # dot = False significa o contrario, exemplo IPs que usam ponto na string
    if ';' in string:
        delimitador = ";"
    elif ',' in string:
        delimitador = ","
    elif '.' in string:
        if dot:
            delimitador = "."
        else:
            delimitador = " "
    else:
        delimitador = " "
    # Processando e imprimindo string
    lista = list(string.split(delimitador))
    for itens in lista:
        print(esp * 3 + "\"" + itens.strip() + "\",")
        # OBS nao funcionou usando o return na funcao


def main():

    wbA = openpyxl.load_workbook(arqA, data_only=True)  # Abrindo arquivo
    sheetA = wbA.get_sheet_by_name(abaA)  # Carregando aba do arquivo

    # Convertendo letras de colunas em numeros
    colunasAn = []
    for letra in colunasA:
        colunasAn.append(column_index_from_string(letra))

    # Percorrendo planilha para gerar terraform
    for row in range(inicioA, fimA):
        print(esp * 1 + "rule {")
        print(esp * 2 + "name = \"" + str(sheetA.cell(row=row, column=colunasAn[0]).value) + "\"\n")
        print(esp * 2 + "source_addresses = [")
        proc_string(str(sheetA.cell(row=row, column=colunasAn[1]).value), dot=False)
        print(esp * 2 + "]" + "\n")
        print(esp * 2 + "destination_ports = [")
        proc_string(str(sheetA.cell(row=row, column=colunasAn[2]).value), dot=True)
        print(esp * 2 + "]" + "\n")
        print(esp * 2 + "destination_addresses = [")
        proc_string(str(sheetA.cell(row=row, column=colunasAn[3]).value), dot=False)
        print(esp * 2 + "]" + "\n")
        print(esp * 2 + "protocols = [")
        proc_string(str(sheetA.cell(row=row, column=colunasAn[4]).value), dot=True)
        print(esp * 2 + "]")
        print(esp * 1 + "}\n")
    print('}')  # Finalizando a collection


if __name__ == '__main__':
    main()
