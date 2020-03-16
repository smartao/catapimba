#!/usr/bin/python3
import openpyxl
from openpyxl.utils import column_index_from_string

# ######### Variavies de configuracoes #########

# Arquivo que contem o grupo maior de elementos
# com todos os elementos que DEVEM permanacer
arqA = "srv-novos.xlsx"
abaA = "Aplicacoes"
colMatchA = "C"
inicioA = 6  # Numero da linha que os dados iniciam

# Arquivo que contem o grupo de elementos ja editados
# com os elementos que seram removidos.
# Se os elementos existirem os dados já preenchidos seram copiados
arqB = "srv-de-para.xlsx"
abaB = "Servidores"
colMatchB = "B"
inicioB = 8  # Numero da linha que os dados iniciam

arqMerge = "merge.xlsx"  # Arquivo que tera os dados mesclados das planilhas
abaMerge = "ServidoresMerge"  # Aba que recebera os dados mesclados

# ##############################################

# Abrindo arquivos
wbA = openpyxl.load_workbook(arqA)
wbB = openpyxl.load_workbook(arqB)
wbC = openpyxl.load_workbook(arqMerge)

# Abrind abas das planilhas
sheetA = wbA.get_sheet_by_name(abaA)
sheetB = wbB.get_sheet_by_name(abaB)
sheetMerge = wbC.get_sheet_by_name(abaMerge)

# Listas que serao utilizadas
srvA = []
srvB = []
# linhaB = []

# Percorrendo planilhaA e armazenando todos os valores em srvA
for row in range(inicioA, sheetA.max_row + 1):
    srvA.append(sheetA[colMatchA + str(row)].value)

# Percorrendo planilhaB e armazenando todos os valores em srvB
for row in range(inicioB, sheetB.max_row + 1):
    srvB.append(sheetB[colMatchB + str(row)].value)
    # linhaB.append(row)  # Armazenando o numero da linha que o elemento esta

# Limpando previamente todos os dados da planilha de merge
for row in sheetMerge.rows:
    for cell in row:
        sheetMerge.cell(row=cell.row, column=cell.column).value = ""

# Comcecando da linha InicioB,
# para assim manter a mesma formatacao da planilhaB
a = inicioB

for srv in srvA:  # Validando se o elemento srv existe dentro da lista srv1
    if srv in srvB:  # Caso existir pegue o elemento, item COMUM
        index = srvB.index(srv)  # Armazenando o index do elemento
        # Percorrendo todas as colunas do arquivo
        for col in range(1, sheetB.max_column + 1):
            # Gravando todos valores das colunas do arquivoB no arquivo Merge
            sheetMerge.cell(row=a, column=col).value = sheetB.cell(
                row=index+inicioB, column=col).value
        # Gravando na ultima coluna o valor comum
        sheetMerge.cell(row=a, column=sheetB.max_column+1).value = "Comum"
    else:  # Caso o elemento nao exista, item INCOMUM
        # Caso o item seja incomum, gravar apenas o nome do servidor
        sheetMerge.cell(row=a, column=column_index_from_string(
            colMatchB)).value = srv
        # Gravando na ultima coluna o valor incomum
        sheetMerge.cell(row=a, column=sheetB.max_column+1).value = "Incomum"
    a = a + 1

# Percorrendo novamente todas as colunas do arB para gravar o cabecalho
for col in range(1, sheetB.max_column + 1):
    # Gravando os valores do cabecalho da primeira linha
    # inicioB-1 = Valor do cabecalho da planilha B
    sheetMerge.cell(
        row=inicioB-1, column=col).value = sheetB.cell(row=inicioB-1, column=col).value

sheetMerge.cell(row=inicioB-1, column=sheetB.max_column+1).value = "Tipo"

# Salva as alteracoes no arquivo de merge
wbC.save(arqMerge)
