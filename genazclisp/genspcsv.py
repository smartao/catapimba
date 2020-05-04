#!/usr/bin/python3
import openpyxl
from openpyxl.utils import column_index_from_string

# ######### Variavies de configuracoes #########

arqA = "lista.xlsx"
abaA = "Service principals"
inicioA = 5  # Inicio da linha com os dados
# status, role, idapp, displayname, homepage, logouturl, scope
colunasA = ["A", "C", "D", "F", "G", "H", "N"]
dirlogs = "logs"
statusnaodesejado = "NãoVai"

# ##########################################

def main():
    wbA = openpyxl.load_workbook(arqA, data_only=True)  # Abrindo arquivo
    sheetA = wbA.get_sheet_by_name(abaA)  # Carregando aba do arquivo
    fimA = sheetA.max_row  # ultimo linha da planilha
    
    # Convertendo letras de colunas em numeros
    colunasAn = []
    for letra in colunasA:
        colunasAn.append(column_index_from_string(letra))

    # Percorrendo planilha para gerar terraform
    for row in range(inicioA, fimA+1):
        
        status = sheetA.cell(row=row, column=colunasAn[0]).value
        
        # Validando o status na coluna A  
        if status != statusnaodesejado:
            role = sheetA.cell(row=row, column=colunasAn[1]).value
            idapp = sheetA.cell(row=row, column=colunasAn[2]).value
            
            sp = sheetA.cell(row=row, column=colunasAn[3]).value
            spnew = sp.replace(" ", "-")
            
            homepage = sheetA.cell(row=row, column=colunasAn[4]).value
            logouturl = sheetA.cell(row=row, column=colunasAn[5]).value
            scope = sheetA.cell(row=row, column=colunasAn[6]).value
            
            # Imprimindo reusltando na tela
            print (idapp + "," + role + "," + spnew + "," + str(homepage) + "," + str(logouturl) + "," + str(scope))


if __name__ == '__main__':
    main()
