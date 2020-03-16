#!/usr/bin/python3
import openpyxl
from openpyxl.utils import column_index_from_string

'''
Merge simplesAB
O merge que essa script faz e de 1 elemento para um elemento

Objetivo
Comparar todos os elementos da aba "abaA" na coluna "colMatchA"
com os elementos da aba "abaB" da coluna "colMatchB"
Quando ocorrer o match, pegar o valor da coluna "colTarget"
e salvar na coluna "colresult" da aba "abaB"
'''

# ####### Variavies de configuracoes #######

arq = "servidores-exemplo.xlsx"  # Arquivo de exel para ser analizado

abaA = "Servidores"  # Aba que contem o grupo de dados maior
colMatchA = "A"  # Coluna que os dados seram comparados
inicioA = 8  # Numero da linha que os dados iniciam

abaB = "Aplicacoes"  # Aba que contem o grupo de dados menor
colMatchB = "B"  # Coluna
inicioB = 6  # Numero da linha que os dados iniciam

# Coluna que os elmentos sera coletados
# Para salvar na planilha
colTarget = "A"
colresult = "H"  # Coluna que sera salvo o resultado

# ##########################################

# Abrindo o arquivo Excel
wb1 = openpyxl.load_workbook(arq)

# Abrindo as planilhas de servidores e aplicacoes
sheetSrv = wb1.get_sheet_by_name(abaA)
sheetApp = wb1.get_sheet_by_name(abaB)

srv1 = []  # Lista para armazenar os servidores
srv2 = []  # Lista para armazenar os produtos e pegar os codigos
target = []  # Codigos dos repectivos produtos

# Percorrendo aba com servidores
for row in range(inicioA, sheetSrv.max_row + 1):
    srv1.append(sheetSrv[colMatchA + str(row)].value)

# Percorrendo planilha que tem servidores e aplicações
for row in range(inicioB, sheetApp.max_row + 1):
    srv2.append(sheetApp[colMatchB + str(row)].value)
    target.append(sheetApp[colTarget + str(row)].value)

# Nova coluna onde sera granva os dados
novacoluna = column_index_from_string(colresult)

# Percorrendo todos os elementos da lista srv1
for i in range(len(srv1)):
    # Se o objeto existir na dentro da lista srv2
    if srv1[i] in srv2:
        # Pega o index do objeto
        index = srv2.index(srv1[i])
        # Armazenando o valor do taget em resulta
        result = target[index]
    else:
        result = 'SRV NAO IDENTIFICADO'
    # Gravando os valores do campo
    sheetSrv.cell(row=inicioA+i, column=novacoluna).value = result

# Gravando o cabecalho do campo
sheetSrv.cell(row=inicioA-1, column=novacoluna).value = abaB

wb1.save(arq)  # Salvando as modificacoes no arquivo
